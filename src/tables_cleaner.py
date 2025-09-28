from .config import *

import os
import glob

import openpyxl


def clean(table):
    wb = openpyxl.load_workbook(table)
    table_name = os.path.basename(table)

    for sheetname in wb.sheetnames:

        ws = wb[sheetname]
        print(f"Processing: {sheetname}")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if cell.column_letter != 'B':
                    if cell.font.strike or (isinstance(cell.value, str) and '.' not in cell.value) or cell.value =='—':
                        cell.value = None

    wb.save(os.path.join(CLEAN_PATH, table_name))

    print(f'{table_name}: Successfully cleaned')



def clean_all_tables():
    tables = [os.path.join(RAW_PATH, f"{name}_2024.xlsx") for name in COMPETITIONS_NAMES]

    for table in tables:
        if os.path.exists(table):
            clean(table)
        else:
            print("Incorrect raw table path:", table)


def main():
    clean_all_tables()


if __name__ == "__main__":
    main()